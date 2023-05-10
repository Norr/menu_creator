import json
from openpyxl import load_workbook
import os
import base64
import jinja2
import pdfkit
from pathlib import Path

POINT = 0.3527777778


config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')

with open(file='de_translate.json', mode='r', encoding='utf-8') as translations:
    de_client_tranlations = json.loads(translations.read())

def menu_to_dict(filename:str="menu.xlsx") -> dict:
    """Function returns dict of excel file contains menu data.

    :param filename: Excel file contains menu data.
    :type filename: str
    :return: Dict of data.
    :rtype: dict
    """

    workbook = load_workbook(filename=filename)
    sections = workbook.sheetnames
    menu_data = {}
    for section in sections:
        menu_data[section] = tuple(element for element in workbook[section].values)
    return menu_data

def image_dict(dir:str='images')->dict:
    """Function that create dict of images, where key is file name and valu is base64 encode representation of image.

    :param dir: path to folder with images, defaults to 'images'
    :type dir: str, optional
    :return: dict with pair file name: base64 file representation
    :rtype: dict
    """
    dict_of_img = {}
    for file in os.listdir(dir):
        with open(file=os.path.join(dir, file), mode='rb') as image:
            dict_of_img[file] = base64.b64encode(image.read()).decode('utf-8')
    return dict_of_img         

def _html2pdf(html_file, output_file, **options):
    """
    Function that create pdf file from rendered html template.
    :return: PDF file
    """
    with open(file=html_file) as f:
        pdfkit.from_file(f, output_file, options=options, configuration=config, css=os.path.join(os.getcwd(), 'html_template', 'style.css'))
#test

def render_pdf(format:str='A4'):
        """
        Method that creating html page with email contents like Microsoft Outlook print style.
        :return: file rendered.html
        """
        template_loader = jinja2.FileSystemLoader(searchpath=os.path.join(os.getcwd()))
        template_env = jinja2.Environment(loader=template_loader)
        template_file = os.path.join(os.getcwd(), 'html_template', 'template.html')
        rendered_file = 'rendered.html'
        template = template_env.get_template(template_file)
        output_data = template.render(menu_to_dict(filename='menu.xlsx'))
        with open(rendered_file, mode='w', encoding='utf8') as rendered:
            rendered.write(output_data)
        _html2pdf(rendered_file, f"menu_{format}.pdf")

